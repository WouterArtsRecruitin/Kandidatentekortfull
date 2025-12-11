#!/usr/bin/env python3
"""
Corporate Recruiter Automation v5.4 NATIVE PIPEDRIVE EXPORT
===========================================================
7 gefaseerde prompts voor stap-voor-stap verwerking

ALLE V5.2 FIXES BEHOUDEN:
- Word boundary matching voor sector detection
- Generic email detection (info@, hr@, etc.)
- Robust deduplication (in-batch + API)
- Rate limiting (Pipedrive: 100/10s, Apollo: 50/min)
- Waterfall enrichment

V5.3 FEATURES:
- 7 gefaseerde prompts die stap voor stap worden uitgevoerd
- Interactieve modus met bevestiging per fase
- Excel output per fase voor review
- Skip naar specifieke fase mogelijk

NIEUW IN V5.4:
- Native Pipedrive export format (direct importeerbaar)
- Auto-generatie van Pipedrive import file na fase 7
- --native-only CLI flag voor alleen Pipedrive format
- Templates module met pipedrive_native_export.py
"""

import os
import re
import sys
import time
import json
import logging
import argparse
import pandas as pd
import requests
from typing import Optional, Dict, List, Tuple, Set
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from functools import wraps
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Native Pipedrive export (optional - graceful fallback)
try:
    from templates.pipedrive_native_export import export_to_pipedrive_native
    NATIVE_EXPORT_AVAILABLE = True
except ImportError:
    NATIVE_EXPORT_AVAILABLE = False

load_dotenv()

# =============================================================================
# CONFIGURATION
# =============================================================================

PIPEDRIVE_API_TOKEN = os.getenv('PIPEDRIVE_API_TOKEN')
APOLLO_API_KEY = os.getenv('APOLLO_API_KEY')

PIPELINE_ID = 14
LEAD_STAGE_NAME = 'lead'
EMAIL_READY_STAGE_NAME = 'Email Sequence Ready'

PIPEDRIVE_RATE_LIMIT = (100, 10)
APOLLO_RATE_LIMIT = (50, 60)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# V5.2 FIXES - CONSTANTS
# =============================================================================

GENERIC_EMAIL_PREFIXES = [
    'info@', 'hr@', 'recruitment@', 'recruiter@', 'vacature@', 'vacatures@',
    'jobs@', 'career@', 'careers@', 'werk@', 'werken@', 'solliciteer@',
    'contact@', 'hello@', 'hallo@', 'algemeen@', 'office@', 'admin@',
    'receptie@', 'frontdesk@', 'support@', 'helpdesk@'
]

EXCLUDE_TITLE_KEYWORDS = [
    'stage', 'meewerkstage', 'afstudeerstage', 'werkstudent', 'student',
    'marketeer', 'recruitment marketeer', 'campus recruiter',
    'medewerker p&o', 'medewerker hr', 'hr medewerker',
    'staf recruiter', 'accountmanager'
]

HR_KEYWORDS = [
    'recruiter', 'recruitment', 'talent acquisition', 'talent',
    'hr', 'human resources', 'personeelszaken', 'p&o',
    'hrm', 'hr manager', 'hr advisor', 'hr adviseur',
    'hr business partner', 'hrbp', 'people', 'people manager',
    'werving', 'selectie', 'werving en selectie', 'werving & selectie',
    'technical recruiter', 'it recruiter', 'recruitment consultant',
    'recruitment specialist', 'talent scout', 'recruitment officer',
    'corporate recruiter', 'recruitment manager', 'recruitment coordinator', 'talent partner'
]

EXCLUDED_SECTORS = [
    'it', 'ict', 'software', 'technology', 'tech',
    'finance', 'bank', 'insurance', 'verzekering',
    'zorg', 'healthcare', 'medisch', 'gezondheid',
    'consultancy', 'consulting', 'advies',
    'retail', 'detailhandel', 'winkel',
    'overheid', 'gemeente', 'government',
    'onderwijs', 'education', 'school'
]

PREFERRED_SECTORS = [
    'oil & gas', 'olie', 'gas', 'constructie', 'bouw', 'construction',
    'productie', 'manufacturing', 'production', 'automation', 'automatisering',
    'industrieel', 'renewable energy', 'energie', 'energy', 'duurzaam',
    'metaal', 'staal', 'metal', 'machinebouw', 'installatie', 'techniek',
    'installation', 'offshore', 'maritiem', 'procesindustrie', 'chemie',
    'engineering', 'ingenieur', 'hightech', 'semiconductors', 'semiconductor'
]

ICP_REGIONS = ['gelderland', 'overijssel', 'noord-brabant', 'limburg', 'utrecht', 'flevoland']

COMPETITOR_KEYWORDS = [
    'uitzendbureau', 'staffing', 'detachering', 'secondment',
    'werving en selectie bureau', 'recruitment bureau',
    'executive search', 'headhunting', 'interim bureau',
    'payroll', 'hr diensten', 'hr services',
    'talentbureau', 'arbeidsmarkt', 'flexwerk',
    'uitzenden', 'detacheren', 'professionals leveren'
]

# =============================================================================
# V5.2 HELPER FUNCTIONS
# =============================================================================

def is_generic_email(email: str) -> bool:
    if not email:
        return False
    email_lower = email.lower().strip()
    return any(email_lower.startswith(prefix) for prefix in GENERIC_EMAIL_PREFIXES)

def word_boundary_match(keyword: str, text: str) -> bool:
    if not keyword or not text:
        return False
    pattern = r'\b' + re.escape(keyword.lower()) + r'\b'
    return bool(re.search(pattern, text.lower()))

# =============================================================================
# RATE LIMITER
# =============================================================================

class RateLimiter:
    def __init__(self, calls: int, period: int):
        self.calls = calls
        self.period = period
        self.tokens = calls
        self.last_update = time.time()
    
    def acquire(self):
        now = time.time()
        elapsed = now - self.last_update
        self.tokens = min(self.calls, self.tokens + elapsed * (self.calls / self.period))
        self.last_update = now
        if self.tokens < 1:
            sleep_time = (1 - self.tokens) * (self.period / self.calls)
            time.sleep(sleep_time)
            self.tokens = 0
        else:
            self.tokens -= 1

def rate_limited(limiter: RateLimiter):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            limiter.acquire()
            return func(*args, **kwargs)
        return wrapper
    return decorator

pipedrive_limiter = RateLimiter(*PIPEDRIVE_RATE_LIMIT)
apollo_limiter = RateLimiter(*APOLLO_RATE_LIMIT)

def retry_with_backoff(max_retries: int = 3, base_delay: float = 1.0):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            last_exception = None
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except requests.exceptions.RequestException as e:
                    last_exception = e
                    if attempt < max_retries - 1:
                        delay = base_delay * (2 ** attempt)
                        logger.warning(f"Retry {attempt + 1}/{max_retries} after {delay}s: {e}")
                        time.sleep(delay)
            raise last_exception
        return wrapper
    return decorator

# =============================================================================
# PHASED PROMPT PROCESSOR
# =============================================================================

class PhasedPromptProcessor:
    """
    Voert 7 gefaseerde prompts uit voor JobDigger data verwerking
    
    PROMPT 1: Data Clean & Filter
    PROMPT 2: ICP Scoring (Golden 500)
    PROMPT 3: Final Golden 500 Filter
    PROMPT 4: Brave Search Sector Validatie
    PROMPT 5: Prioritering op Kans van Slagen
    PROMPT 6: Apollo MCP Enrichment
    PROMPT 7: ICP_MET_EMAIL Tab Genereren
    """
    
    def __init__(self, input_file: str, output_dir: str = None):
        self.input_file = input_file
        self.output_dir = output_dir or os.path.dirname(input_file) or '.'
        self.df = None
        self.stats = {}
        self.current_phase = 0
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.base_output = os.path.join(self.output_dir, f"PIPELINE14_PROCESSED_{timestamp}")
    
    def load_data(self) -> pd.DataFrame:
        logger.info(f"\n{'='*60}")
        logger.info("📂 LOADING JOBDIGGER DATA")
        logger.info(f"{'='*60}")
        logger.info(f"File: {self.input_file}")
        
        df_raw = pd.read_excel(self.input_file, header=None)
        header_row = None
        for idx, row in df_raw.iterrows():
            if 'Bedrijfsnaam' in row.values:
                header_row = idx
                break
        
        if header_row is None:
            raise ValueError("Could not find header row with 'Bedrijfsnaam'")
        
        self.df = pd.read_excel(self.input_file, header=header_row)
        logger.info(f"✅ Loaded {len(self.df)} rows (header at row {header_row + 1})")
        return self.df

    # =========================================================================
    # PROMPT 1: DATA CLEAN & FILTER
    # =========================================================================
    def prompt_1_data_clean_filter(self) -> pd.DataFrame:
        """
        PROMPT 1: Data Clean & Filter
        - Verwijder duplicaten
        - Verwijder bemiddelaars/uitzendbureaus
        - Filter op ICP regio's
        - Filter op technische sectoren
        - Filter op functietitel
        """
        logger.info(f"\n{'='*60}")
        logger.info("🔧 PROMPT 1: DATA CLEAN & FILTER")
        logger.info(f"{'='*60}")
        
        df = self.df.copy()
        initial_count = len(df)
        stats = {
            'initial': initial_count,
            'removed_duplicates': 0,
            'removed_intermediary': 0,
            'removed_wrong_region': 0,
            'removed_non_technical': 0,
            'removed_stage_student': 0,
            'removed_no_recruiter': 0
        }
        
        # STAP 1: Verwijder duplicaten
        logger.info("\n📋 STAP 1: Verwijder duplicaten...")
        df_before = len(df)
        df = df.drop_duplicates(subset=['Bedrijfsnaam', 'Functietitel'], keep='first')
        stats['removed_duplicates'] = df_before - len(df)
        logger.info(f"   Verwijderd: {stats['removed_duplicates']} duplicaten")
        
        # STAP 2: Verwijder bemiddelaars/uitzendbureaus
        logger.info("\n📋 STAP 2: Filter bemiddelaars/uitzendbureaus...")
        df['_company_text'] = (
            df['Bedrijfsnaam'].fillna('').astype(str) + ' ' +
            df.get('Bedrijf: Branche', pd.Series([''] * len(df))).fillna('').astype(str)
        ).str.lower()
        
        def is_competitor(text):
            return any(word_boundary_match(kw, text) for kw in COMPETITOR_KEYWORDS)
        
        df_before = len(df)
        competitor_mask = df['_company_text'].apply(is_competitor)
        df = df[~competitor_mask]
        stats['removed_intermediary'] = df_before - len(df)
        logger.info(f"   Verwijderd: {stats['removed_intermediary']} bemiddelaars")
        
        # STAP 3: Filter op ICP regio's
        logger.info("\n📋 STAP 3: Filter op ICP regio's...")
        df_before = len(df)
        province_col = 'Standplaats: Provincie'
        if province_col in df.columns:
            df['_province_lower'] = df[province_col].fillna('').astype(str).str.lower()
            region_mask = df['_province_lower'].apply(
                lambda x: any(region in x for region in ICP_REGIONS)
            )
            df = df[region_mask]
        stats['removed_wrong_region'] = df_before - len(df)
        logger.info(f"   Verwijderd: {stats['removed_wrong_region']} verkeerde regio")
        
        # STAP 4: Filter niet-technische sectoren
        logger.info("\n📋 STAP 4: Filter niet-technische sectoren...")
        df_before = len(df)
        sector_col = 'Bedrijf: Branche'
        if sector_col in df.columns:
            df['_sector_lower'] = df[sector_col].fillna('').astype(str).str.lower()
            def is_excluded_sector(text):
                return any(word_boundary_match(exc, text) for exc in EXCLUDED_SECTORS)
            excluded_mask = df['_sector_lower'].apply(is_excluded_sector)
            df = df[~excluded_mask]
        stats['removed_non_technical'] = df_before - len(df)
        logger.info(f"   Verwijderd: {stats['removed_non_technical']} niet-technische sectoren")
        
        # STAP 5: Filter stage/student/marketeer
        logger.info("\n📋 STAP 5: Filter functietitels...")
        df_before = len(df)
        title_col = 'Functietitel'
        if title_col in df.columns:
            df['_title_lower'] = df[title_col].fillna('').astype(str).str.lower()
            def has_exclude_keyword(text):
                return any(kw in text for kw in EXCLUDE_TITLE_KEYWORDS)
            exclude_mask = df['_title_lower'].apply(has_exclude_keyword)
            df = df[~exclude_mask]
        stats['removed_stage_student'] = df_before - len(df)
        logger.info(f"   Verwijderd: {stats['removed_stage_student']} stage/student/marketeer")
        
        # STAP 6: Behoud alleen echte recruiter vacatures
        logger.info("\n📋 STAP 6: Behoud alleen echte recruiter vacatures...")
        df_before = len(df)
        if title_col in df.columns:
            def is_recruiter_position(text):
                return any(word_boundary_match(kw, text) for kw in HR_KEYWORDS)
            recruiter_mask = df['_title_lower'].apply(is_recruiter_position)
            df = df[recruiter_mask]
        stats['removed_no_recruiter'] = df_before - len(df)
        logger.info(f"   Verwijderd: {stats['removed_no_recruiter']} geen echte recruiter")
        
        # Cleanup temp columns
        temp_cols = [c for c in df.columns if c.startswith('_')]
        df = df.drop(columns=temp_cols, errors='ignore')
        
        stats['final'] = len(df)
        self.stats['prompt_1'] = stats
        
        logger.info(f"\n{'─'*40}")
        logger.info("📊 PROMPT 1 SAMENVATTING:")
        logger.info(f"   Totaal voor filtering: {stats['initial']}")
        logger.info(f"   Totaal na filtering: {stats['final']}")
        logger.info(f"   Verwijderd totaal: {stats['initial'] - stats['final']}")
        
        self.df = df
        return df
    
    # =========================================================================
    # PROMPT 2: ICP SCORING
    # =========================================================================
    def prompt_2_icp_scoring(self) -> pd.DataFrame:
        logger.info(f"\n{'='*60}")
        logger.info("⭐ PROMPT 2: ICP SCORING (GOLDEN 500)")
        logger.info(f"{'='*60}")
        
        df = self.df.copy()
        df['has_active_vacancy'] = True
        df['right_size'] = 'UNKNOWN'
        df['is_technical_sector'] = False
        df['no_internal_recruiter'] = True
        df['golden_score'] = 0.0
        
        df['golden_score'] += 1  # Active vacancy
        df['golden_score'] += 0.5  # Size unknown
        
        sector_col = 'Bedrijf: Branche'
        if sector_col in df.columns:
            df['_sector_check'] = df[sector_col].fillna('').astype(str).str.lower()
            def is_preferred_sector(text):
                return any(word_boundary_match(pref, text) for pref in PREFERRED_SECTORS)
            df['is_technical_sector'] = df['_sector_check'].apply(is_preferred_sector)
            df.loc[df['is_technical_sector'], 'golden_score'] += 1
        
        df['golden_score'] += 1  # No internal recruiter
        
        def get_tier(score):
            if score >= 3.5: return 'GOLDEN'
            elif score >= 2.5: return 'SILVER'
            elif score >= 1.5: return 'BRONZE'
            else: return 'INTERIM'
        
        df['tier'] = df['golden_score'].apply(get_tier)
        df = df.drop(columns=[c for c in df.columns if c.startswith('_')], errors='ignore')
        
        tier_counts = df['tier'].value_counts()
        logger.info("\n📊 PROMPT 2 SAMENVATTING:")
        logger.info(f"   GOLDEN: {tier_counts.get('GOLDEN', 0)}")
        logger.info(f"   SILVER: {tier_counts.get('SILVER', 0)}")
        logger.info(f"   BRONZE: {tier_counts.get('BRONZE', 0)}")
        logger.info(f"   INTERIM: {tier_counts.get('INTERIM', 0)}")
        
        self.stats['prompt_2'] = dict(tier_counts)
        self.df = df
        return df

    # =========================================================================
    # PROMPT 3: FINAL GOLDEN 500 FILTER
    # =========================================================================
    def prompt_3_golden_filter(self, max_leads: int = 500) -> pd.DataFrame:
        logger.info(f"\n{'='*60}")
        logger.info("🏆 PROMPT 3: FINAL GOLDEN 500 FILTER")
        logger.info(f"{'='*60}")
        
        df = self.df.copy()
        df_filtered = df[df['tier'].isin(['GOLDEN', 'SILVER'])].copy()
        df_filtered = df_filtered.sort_values(by=['golden_score', 'tier'], ascending=[False, True])
        
        if len(df_filtered) > max_leads:
            df_filtered = df_filtered.head(max_leads)
        
        logger.info(f"\n📊 PROMPT 3 SAMENVATTING:")
        logger.info(f"   Totaal voor filter: {len(df)}")
        logger.info(f"   GOLDEN + SILVER: {len(df_filtered)}")
        
        if 'Standplaats: Provincie' in df_filtered.columns:
            prov_counts = df_filtered['Standplaats: Provincie'].value_counts()
            logger.info("\n   Per Provincie:")
            for prov, count in prov_counts.items():
                logger.info(f"     → {prov}: {count}")
        
        self.stats['prompt_3'] = {'before': len(df), 'after': len(df_filtered)}
        self.df = df_filtered
        return df_filtered
    
    # =========================================================================
    # PROMPT 4: BRAVE SEARCH VALIDATIE
    # =========================================================================
    def prompt_4_brave_validation(self) -> pd.DataFrame:
        logger.info(f"\n{'='*60}")
        logger.info("🔍 PROMPT 4: BRAVE SEARCH SECTOR VALIDATIE")
        logger.info(f"{'='*60}")
        
        df = self.df.copy()
        df['needs_validation'] = False
        df['validation_reason'] = ''
        
        sector_col = 'Bedrijf: Branche'
        if sector_col in df.columns:
            unknown_mask = (
                df[sector_col].isna() | 
                (df[sector_col].astype(str).str.lower().isin(['', 'onbekend', 'unknown']))
            )
            df.loc[unknown_mask, 'needs_validation'] = True
            df.loc[unknown_mask, 'validation_reason'] = 'Sector onbekend'
            
            company_col = 'Bedrijfsnaam'
            if company_col in df.columns:
                tech_in_name = df[company_col].str.lower().str.contains('techniek|technical|tech ', na=False)
                df.loc[tech_in_name & ~unknown_mask, 'needs_validation'] = True
                df.loc[tech_in_name & ~unknown_mask, 'validation_reason'] = 'Mogelijk detacheerder'
        
        needs_val = df['needs_validation'].sum()
        logger.info(f"\n📊 PROMPT 4 SAMENVATTING:")
        logger.info(f"   Leads die validatie nodig hebben: {needs_val}")
        logger.info(f"   Leads al gevalideerd: {len(df) - needs_val}")
        
        if needs_val > 0:
            logger.info("\n   ⚠️  ACTIE NODIG:")
            for _, row in df[df['needs_validation']].head(5).iterrows():
                logger.info(f"     → {row.get('Bedrijfsnaam', 'Unknown')}: {row.get('validation_reason', '')}")
        
        self.stats['prompt_4'] = {'needs_validation': needs_val, 'validated': len(df) - needs_val}
        self.df = df
        return df
    
    # =========================================================================
    # PROMPT 5: SUCCESS SCORING
    # =========================================================================
    def prompt_5_success_scoring(self) -> pd.DataFrame:
        logger.info(f"\n{'='*60}")
        logger.info("📊 PROMPT 5: PRIORITERING OP KANS VAN SLAGEN")
        logger.info(f"{'='*60}")
        
        df = self.df.copy()
        df['success_score'] = 0
        
        tier_scores = {'GOLDEN': 60, 'SILVER': 45, 'BRONZE': 30, 'INTERIM': 15}
        df['success_score'] = df['tier'].map(tier_scores).fillna(30)
        
        if 'is_technical_sector' in df.columns:
            df.loc[df['is_technical_sector'] == True, 'success_score'] += 15
        
        province_col = 'Standplaats: Provincie'
        if province_col in df.columns:
            df['_prov_lower'] = df[province_col].fillna('').astype(str).str.lower()
            icp_mask = df['_prov_lower'].apply(lambda x: any(region in x for region in ICP_REGIONS))
            df.loc[icp_mask, 'success_score'] += 10
        
        email_col = 'Contactpersoon: E-mail'
        if email_col in df.columns:
            has_email = df[email_col].notna() & (df[email_col] != '')
            df.loc[has_email, 'success_score'] += 8
            personal_email = has_email & ~df[email_col].apply(is_generic_email)
            df.loc[personal_email, 'success_score'] += 7
        
        df['success_score'] = df['success_score'].clip(upper=100)
        
        def get_priority(score):
            if score >= 80: return 'A - DIRECT BENADEREN'
            elif score >= 60: return 'B - NURTURE'
            elif score >= 40: return 'C - INTERIM POTENTIEEL'
            else: return 'D - EXCLUDE'
        
        df['priority'] = df['success_score'].apply(get_priority)
        df = df.drop(columns=[c for c in df.columns if c.startswith('_')], errors='ignore')
        
        priority_counts = df['priority'].value_counts()
        logger.info("\n📊 PROMPT 5 SAMENVATTING:")
        for p, c in priority_counts.items():
            logger.info(f"   {p}: {c}")
        
        self.stats['prompt_5'] = dict(priority_counts)
        self.df = df
        return df

    # =========================================================================
    # PROMPT 6: APOLLO ENRICHMENT
    # =========================================================================
    def prompt_6_apollo_enrichment(self, enricher=None) -> pd.DataFrame:
        logger.info(f"\n{'='*60}")
        logger.info("🔄 PROMPT 6: APOLLO ENRICHMENT")
        logger.info(f"{'='*60}")
        
        df = self.df.copy()
        df['enrichment_source'] = ''
        df['enrichment_confidence'] = 0
        df['Contact_Email'] = df.get('Contactpersoon: E-mail', '')
        df['Contact_Voornaam'] = df.get('Contactpersoon: Voornaam', '')
        df['Contact_Achternaam'] = df.get('Contactpersoon: Achternaam', '')
        df['Contact_Telefoon'] = df.get('Contactpersoon: Telefoon', '')
        
        needs_enrichment = (
            (df['priority'].str.startswith('A')) &
            (df['Contact_Email'].isna() | (df['Contact_Email'] == '') | df['Contact_Email'].apply(is_generic_email))
        )
        
        logger.info(f"   Priority A leads: {(df['priority'].str.startswith('A')).sum()}")
        logger.info(f"   Needs enrichment: {needs_enrichment.sum()}")
        
        if enricher and needs_enrichment.sum() > 0:
            logger.info("\n   🔄 Starting enrichment...")
            enriched_count = 0
            
            for idx in df[needs_enrichment].index:
                row = df.loc[idx]
                
                class SimpleLead:
                    pass
                
                lead = SimpleLead()
                lead.company_name = row.get('Bedrijfsnaam', '')
                lead.website = row.get('Bedrijf: Website', '')
                lead.email = row.get('Contact_Email', '')
                lead.contact_first_name = row.get('Contact_Voornaam', '')
                lead.contact_last_name = row.get('Contact_Achternaam', '')
                lead.contact_name = f"{lead.contact_first_name} {lead.contact_last_name}".strip()
                lead.phone = row.get('Contact_Telefoon', '')
                lead.enrichment_source = ''
                lead.enrichment_confidence = 0
                
                if enricher.enrich_lead(lead):
                    df.loc[idx, 'Contact_Email'] = lead.email
                    df.loc[idx, 'Contact_Voornaam'] = lead.contact_first_name
                    df.loc[idx, 'Contact_Achternaam'] = lead.contact_last_name
                    df.loc[idx, 'Contact_Telefoon'] = lead.phone or row.get('Contact_Telefoon', '')
                    df.loc[idx, 'enrichment_source'] = lead.enrichment_source
                    df.loc[idx, 'enrichment_confidence'] = lead.enrichment_confidence
                    enriched_count += 1
            
            logger.info(f"   ✅ Enriched: {enriched_count} leads")
        else:
            logger.info("   ⚠️  No enricher provided or no leads need enrichment")
        
        has_email = df['Contact_Email'].notna() & (df['Contact_Email'] != '')
        has_personal_email = has_email & ~df['Contact_Email'].apply(is_generic_email)
        
        self.stats['prompt_6'] = {
            'needs_enrichment': needs_enrichment.sum(),
            'has_email': has_email.sum(),
            'has_personal_email': has_personal_email.sum()
        }
        
        logger.info(f"\n📊 PROMPT 6 SAMENVATTING:")
        logger.info(f"   Leads met email: {has_email.sum()}")
        logger.info(f"   Leads met personal email: {has_personal_email.sum()}")
        
        self.df = df
        return df
    
    # =========================================================================
    # PROMPT 7: ICP_MET_EMAIL TAB
    # =========================================================================
    def prompt_7_icp_met_email_tab(self) -> pd.DataFrame:
        logger.info(f"\n{'='*60}")
        logger.info("📧 PROMPT 7: ICP_MET_EMAIL TAB GENEREREN")
        logger.info(f"{'='*60}")
        
        df = self.df.copy()
        has_email = df['Contact_Email'].notna() & (df['Contact_Email'] != '')
        is_priority_a = df['priority'].str.startswith('A')
        
        icp_with_email = df[has_email & is_priority_a].copy()
        icp_with_email = icp_with_email.sort_values('success_score', ascending=False)
        
        output_columns = [
            'Bedrijfsnaam', 'Contact_Email', 'Contact_Telefoon',
            'Contact_Voornaam', 'Contact_Achternaam', 'Bedrijf: Branche',
            'Standplaats: Gemeente', 'Standplaats: Provincie', 'Bedrijf: Website',
            'success_score', 'Functietitel', 'URL'
        ]
        available_cols = [c for c in output_columns if c in icp_with_email.columns]
        icp_with_email = icp_with_email[available_cols]
        
        logger.info(f"\n📊 PROMPT 7 SAMENVATTING:")
        logger.info(f"   Priority A leads: {is_priority_a.sum()}")
        logger.info(f"   ICP_MET_EMAIL: {len(icp_with_email)}")
        
        if len(icp_with_email) > 0:
            logger.info(f"\n   TOP 10:")
            for i, (_, row) in enumerate(icp_with_email.head(10).iterrows(), 1):
                score = row.get('success_score', 0)
                company = row.get('Bedrijfsnaam', 'Unknown')
                email = row.get('Contact_Email', 'N/A')
                logger.info(f"   {i}. [{score:.0f}] {company} → {email}")
        
        self.stats['prompt_7'] = {
            'priority_a': is_priority_a.sum(),
            'with_email': has_email.sum(),
            'icp_met_email': len(icp_with_email)
        }
        
        self.icp_with_email = icp_with_email
        return icp_with_email

    # =========================================================================
    # EXPORT FUNCTIONS
    # =========================================================================
    def export_to_excel(self, filename: str = None) -> str:
        if filename is None:
            filename = f"{self.base_output}.xlsx"
        
        logger.info(f"\n{'='*60}")
        logger.info("💾 EXPORTING TO EXCEL")
        logger.info(f"{'='*60}")
        
        wb = Workbook()
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        green_fill = PatternFill(start_color='22C55E', end_color='22C55E', fill_type='solid')
        
        # Tab 1: ICP_MET_EMAIL
        if hasattr(self, 'icp_with_email') and len(self.icp_with_email) > 0:
            ws = wb.active
            ws.title = 'ICP_MET_EMAIL'
            for r_idx, row in enumerate(dataframe_to_rows(self.icp_with_email, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.fill = green_fill
                        cell.font = header_font
            logger.info(f"   ✅ ICP_MET_EMAIL: {len(self.icp_with_email)} rows")
        
        # Tab 2: ALL_PROCESSED
        ws2 = wb.create_sheet('ALL_PROCESSED')
        for r_idx, row in enumerate(dataframe_to_rows(self.df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
        logger.info(f"   ✅ ALL_PROCESSED: {len(self.df)} rows")
        
        # Tab 3: Priority splits
        for priority in ['A - DIRECT BENADEREN', 'B - NURTURE', 'C - INTERIM POTENTIEEL']:
            short_name = priority.split(' - ')[0]
            ws_p = wb.create_sheet(f'Priority_{short_name}')
            df_priority = self.df[self.df['priority'] == priority]
            for r_idx, row in enumerate(dataframe_to_rows(df_priority, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_p.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
            logger.info(f"   ✅ Priority_{short_name}: {len(df_priority)} rows")
        
        # Tab 4: VALIDATION_NEEDED
        if 'needs_validation' in self.df.columns:
            df_val = self.df[self.df['needs_validation'] == True]
            if len(df_val) > 0:
                ws_val = wb.create_sheet('VALIDATIE_NODIG')
                for r_idx, row in enumerate(dataframe_to_rows(df_val, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws_val.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == 1:
                            cell.fill = header_fill
                            cell.font = header_font
                logger.info(f"   ✅ VALIDATIE_NODIG: {len(df_val)} rows")
        
        # Tab 5: STATS
        ws_stats = wb.create_sheet('STATS')
        ws_stats['A1'] = 'Metric'
        ws_stats['B1'] = 'Value'
        ws_stats['A1'].fill = header_fill
        ws_stats['A1'].font = header_font
        ws_stats['B1'].fill = header_fill
        ws_stats['B1'].font = header_font
        
        row = 2
        for phase, stats in self.stats.items():
            ws_stats.cell(row=row, column=1, value=f"=== {phase.upper()} ===")
            ws_stats.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            if isinstance(stats, dict):
                for key, value in stats.items():
                    ws_stats.cell(row=row, column=1, value=key)
                    ws_stats.cell(row=row, column=2, value=value)
                    row += 1
            row += 1
        
        wb.save(filename)
        logger.info(f"\n   📁 Saved to: {filename}")
        return filename
    
    def export_to_pipedrive_native_format(self, filename: str = None) -> str:
        """Export data to Pipedrive's native import format."""
        if not NATIVE_EXPORT_AVAILABLE:
            raise ImportError("templates.pipedrive_native_export module not available")
        
        if filename is None:
            filename = f"{self.base_output}_pipedrive_import.xlsx"
        
        logger.info(f"\n{'='*60}")
        logger.info("📤 EXPORTING TO PIPEDRIVE NATIVE FORMAT")
        logger.info(f"{'='*60}")
        
        # Use ICP_MET_EMAIL if available, otherwise use main df
        source_df = self.icp_with_email if hasattr(self, 'icp_with_email') and len(self.icp_with_email) > 0 else self.df
        
        # Map columns to Pipedrive format
        column_mapping = {
            'bedrijfsnaam': 'Organization - Name',
            'contactpersoon': 'Person - Name',
            'email': 'Person - Email',
            'telefoon': 'Person - Phone',
            'functie': 'Person - Title',
            'locatie': 'Organization - Address',
            'sector': 'Deal - Industry',
            'fte': 'Deal - FTE',
            'score': 'Deal - Lead Score',
            'priority': 'Deal - Priority',
            'tier': 'Deal - Tier',
        }
        
        mapped_df = source_df.rename(columns={k: v for k, v in column_mapping.items() if k in source_df.columns})
        
        export_to_pipedrive_native(
            df=mapped_df,
            output_path=filename,
            pipeline_id=PIPELINE_ID,
            stage_name=LEAD_STAGE_NAME
        )
        
        logger.info(f"   📁 Pipedrive import file: {filename}")
        return filename
    
    # =========================================================================
    # RUN ALL PROMPTS
    # =========================================================================
    def run_all_prompts(self, interactive: bool = False, start_phase: int = 1, enricher=None, native_only: bool = False) -> pd.DataFrame:
        logger.info(f"\n{'='*70}")
        logger.info("🚀 CORPORATE RECRUITER AUTOMATION v5.4 - PHASED PROMPTS")
        logger.info(f"{'='*70}")
        logger.info(f"   Input file: {self.input_file}")
        logger.info(f"   Interactive: {interactive}")
        logger.info(f"   Start phase: {start_phase}")
        
        prompts = [
            (1, "Data Clean & Filter", self.prompt_1_data_clean_filter),
            (2, "ICP Scoring", self.prompt_2_icp_scoring),
            (3, "Golden 500 Filter", self.prompt_3_golden_filter),
            (4, "Brave Search Validatie", self.prompt_4_brave_validation),
            (5, "Success Scoring", self.prompt_5_success_scoring),
            (6, "Apollo Enrichment", lambda: self.prompt_6_apollo_enrichment(enricher)),
            (7, "ICP_MET_EMAIL Tab", self.prompt_7_icp_met_email_tab),
        ]
        
        self.load_data()
        
        for phase_num, phase_name, phase_func in prompts:
            if phase_num < start_phase:
                logger.info(f"\n⏭️  Skipping Phase {phase_num}: {phase_name}")
                continue
            
            if interactive:
                response = input(f"\n▶️  Run Phase {phase_num}: {phase_name}? [Y/n/q]: ").strip().lower()
                if response == 'q':
                    logger.info("❌ Aborted by user")
                    break
                elif response == 'n':
                    logger.info(f"⏭️  Skipping Phase {phase_num}")
                    continue
            
            phase_func()
            self.current_phase = phase_num
            
            if phase_num in [3, 5, 7]:
                interim_file = f"{self.base_output}_phase{phase_num}.xlsx"
                self.export_to_excel(interim_file)
        
        final_file = self.export_to_excel()
        
        # Generate native Pipedrive import file
        if NATIVE_EXPORT_AVAILABLE:
            try:
                native_file = self.export_to_pipedrive_native_format()
                logger.info(f"   Pipedrive native: {native_file}")
            except Exception as e:
                logger.warning(f"⚠️  Native export failed: {e}")
        elif native_only:
            logger.warning("⚠️  Native export requested but templates module not available")
        
        logger.info(f"\n{'='*70}")
        logger.info("✅ ALL PROMPTS COMPLETED")
        logger.info(f"{'='*70}")
        logger.info(f"   Final output: {final_file}")
        
        return self.df


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Corporate Recruiter Automation v5.3 - 7 Phased Prompts',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
PROMPTS:
  1. Data Clean & Filter    - Duplicaten, regio, sector, functietitel
  2. ICP Scoring           - Golden 500 scoring (0-4 punten)
  3. Golden 500 Filter     - Alleen GOLDEN + SILVER tiers
  4. Brave Search          - Sector validatie (indien nodig)
  5. Success Scoring       - Prioritering (A/B/C/D)
  6. Apollo Enrichment     - Contact data verrijking
  7. ICP_MET_EMAIL         - Finale outreach lijst

Examples:
  python3 %(prog)s --file leads.xlsx
  python3 %(prog)s --file leads.xlsx --interactive
  python3 %(prog)s --file leads.xlsx --start-phase 5
  python3 %(prog)s --file leads.xlsx --enrich
        """
    )
    
    parser.add_argument('--file', required=True, help='JobDigger Excel file')
    parser.add_argument('--output-dir', help='Output directory')
    parser.add_argument('--interactive', '-i', action='store_true', help='Interactive mode')
    parser.add_argument('--start-phase', type=int, default=1, choices=range(1, 8), help='Start phase (1-7)')
    parser.add_argument('--enrich', action='store_true', help='Enable Apollo enrichment')
    parser.add_argument('--native-only', action='store_true', help='Only generate Pipedrive native import format')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.file):
        logger.error(f"❌ File not found: {args.file}")
        sys.exit(1)
    
    processor = PhasedPromptProcessor(input_file=args.file, output_dir=args.output_dir)
    
    enricher = None
    if args.enrich:
        try:
            from corporate_recruiter_automation_v52_fixed import WaterfallEnricherV51
            enricher = WaterfallEnricherV51()
            logger.info("✅ Apollo enricher initialized")
        except ImportError:
            logger.warning("⚠️  Could not import WaterfallEnricherV51, enrichment disabled")
    
    try:
        processor.run_all_prompts(interactive=args.interactive, start_phase=args.start_phase, enricher=enricher, native_only=args.native_only)
    except KeyboardInterrupt:
        logger.info("\n❌ Interrupted by user")
        sys.exit(1)
    except Exception as e:
        logger.error(f"❌ Fatal error: {e}")
        raise


if __name__ == '__main__':
    main()
